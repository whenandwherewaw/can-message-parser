# -*- coding: utf-8 -*-
"""
CAN 报文解析工具 (GB/T 27930 充电 / GB/T 18487.4 直流V2L放电)
------------------------------------------------------------
导入 CAN 分析仪导出的 CSV 抓包文件, 解析为带颜色区分、逐帧翻译的 Excel 文件。

依赖: Python 3.8+ , openpyxl   (tkinter 为 Python 自带)
安装依赖: pip install openpyxl
"""
import os
import csv
import queue
import threading
import traceback

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

APP_NAME = "CAN 报文解析工具"
APP_VER = "V1.0"

# ============================================================
#                       协议解析核心
# ============================================================

# 多包传输 PGN 名称
PGN_NAME = {
    0x02: 'BRM', 0x06: 'BCP', 0x11: 'BCS',
    0x15: 'BMV', 0x16: 'BMT', 0x17: 'BSP',
    0x31: 'BDR', 0x32: 'ERD',
}

# CAN ID -> (报文代号, 方向)   V=车辆/BMS端发出   D=设备/充电机端发出
ID_MAP = {
    # ---- 车辆(BMS)  -> 设备(充电机/放电设备) ----
    0x1CEC56F4: ('TP.CM', 'V'), 0x1CEB56F4: ('TP.DT', 'V'),
    0x182756F4: ('BHM', 'V'), 0x100956F4: ('BRO', 'V'), 0x181056F4: ('BCL', 'V'),
    0x181356F4: ('BSM', 'V'), 0x181C56F4: ('BSD', 'V'), 0x101956F4: ('BST', 'V'),
    0x181956F4: ('BST(旧ID)', 'V'), 0x181E56F4: ('BEM', 'V'), 0x081E56F4: ('BEM', 'V'),
    0x181556F4: ('BMV', 'V'), 0x181656F4: ('BMT', 'V'), 0x181756F4: ('BSP', 'V'),
    # 18487.4 新增(车端)
    0x183656F4: ('BDC', 'V'), 0x103956F4: ('BDST', 'V'),
    0x183156F4: ('BDR(单帧)', 'V'), 0x183256F4: ('BDR(单帧)', 'V'),
    0x1C3156F4: ('BDR(单帧P7)', 'V'), 0x1C3256F4: ('BDR(单帧P7)', 'V'),
    # ---- 设备(充电机/放电设备) -> 车辆(BMS) ----
    0x1CECF456: ('TP.CM', 'D'), 0x1CEBF456: ('TP.DT', 'D'),
    0x1826F456: ('CHM', 'D'), 0x1801F456: ('CRM', 'D'), 0x1807F456: ('CTS', 'D'),
    0x1808F456: ('CML', 'D'), 0x100AF456: ('CRO', 'D'), 0x1812F456: ('CCS', 'D'),
    0x101AF456: ('CST', 'D'), 0x181DF456: ('CSD', 'D'),
    0x081FF456: ('CEM', 'D'), 0x181FF456: ('CEM', 'D'),
    # 18487.4 新增(设备端)
    0x183DF456: ('ESD', 'D'), 0x103AF456: ('EDST', 'D'),
}


def u16(lo, hi):
    return lo | (hi << 8)


def cur_dis(v):
    """放电电流: 0.1A/位, -400A 偏移 (放电为正)"""
    return v / 10.0 - 400


def cur_chg(v):
    """27930 充电电流: 400 - 0.1A/位"""
    return 400 - v / 10.0


class Decoder:
    """按方向维护 TP 会话状态的解码器"""

    def __init__(self):
        self.sess = {'V': 0, 'D': 0}

    def decode(self, name, drc, d):
        """返回 (解析文本, 显示标签)"""
        try:
            return self._decode(name, drc, d)
        except Exception:
            return '(解析异常)', name

    def _decode(self, name, drc, d):
        # ---------------- 传输协议 ----------------
        if name == 'TP.CM':
            pgn = d[6]
            pn = PGN_NAME.get(pgn, 'PGN%02X' % pgn)
            if d[0] == 0x10:
                self.sess[drc] = pgn
                return 'RTS 请求发送%s: 共%d字节 %d包' % (pn, u16(d[1], d[2]), d[3]), pn + '·RTS'
            if d[0] == 0x11:
                return 'CTS 允许发送%s: 放行%d包, 自第%d包起' % (pn, d[1], d[2]), pn + '·CTS'
            if d[0] == 0x13:
                return 'EndACK %s接收完成(%d字节 %d包)' % (pn, u16(d[1], d[2]), d[3]), pn + '·ACK'
            if d[0] == 0xFF:
                return 'Conn_Abort 放弃%s连接' % pn, pn + '·Abort'
            return '未知TP控制字节 0x%02X' % d[0], name

        if name == 'TP.DT':
            pgn = self.sess.get(drc, 0)
            pn = PGN_NAME.get(pgn, 'PGN%02X' % pgn)
            p = d[0]
            lab = '%s 第%d包' % (pn, p)
            if pn == 'BDR':
                if p == 1:
                    return ('版本V%d.%d, 放电状态bit0-1=%s(01=允许放电), 最大允许放电电流%+.1fA'
                            % (d[1], d[2], format(d[4] & 3, '02b'),
                               cur_dis(u16(d[5], d[6])))), lab
                if p == 2:
                    return ('当前放电电压%.1fV, 最高放电电压%.1fV'
                            % (u16(d[2], d[3]) / 10.0, u16(d[4], d[5]) / 10.0)), lab
            if pn == 'ERD':
                if p == 1:
                    return ('版本V%d.%d, 放电请求bit0-1=%s(01=请求放电), 最小放电电流%+.1fA'
                            % (d[1], d[2], format(d[4] & 3, '02b'),
                               cur_dis(u16(d[5], d[6])))), lab
                if p == 2:
                    lock = d[4] & 3
                    ls = {1: '锁定', 0: '解锁'}.get(lock, format(lock, '02b'))
                    return ('最高放电电压%.1fV, 电子锁状态=%s'
                            % (u16(d[2], d[3]) / 10.0, ls)), lab
            if pn == 'BRM':
                if p == 1:
                    return ('BMS通信协议版本V%d.%d, 电池类型0x%02X, 额定容量%.1fAh'
                            % (d[1], d[2], d[4], u16(d[5], d[6]) / 10.0)), lab
                if p >= 4:
                    return '车辆VIN/BMS软件版本信息', lab
            if pn == 'BCP':
                if p == 1:
                    return ('单体最高允许充电电压%.2fV, 最高允许充电电流%.1fA, 标称总能量%.1fkWh'
                            % (u16(d[1], d[2]) / 100.0, cur_chg(u16(d[3], d[4])),
                               u16(d[5], d[6]) / 10.0)), lab
                if p == 2:
                    # 最高允许充电总电压跨包(第1包B7低字节+本包B1高字节), 此处只报本包可解字段
                    return ('最高允许温度%d℃, 荷电状态SOC=%.1f%%, 当前电池电压%.1fV'
                            % (d[2] - 50, u16(d[3], d[4]) / 10.0,
                               u16(d[5], d[6]) / 10.0)), lab
            if pn == 'BCS':
                if p == 1:
                    return ('充电电压测量值%.1fV, 充电电流测量值%.1fA, SOC=%d%%'
                            % (u16(d[1], d[2]) / 10.0, cur_chg(u16(d[3], d[4])), d[7])), lab
                if p == 2:
                    return '估算剩余充电时间%dmin' % u16(d[1], d[2]), lab
            return '数据包', lab

        # ---------------- 单帧报文 ----------------
        if name == 'CHM':
            return '充电机握手, 通信协议版本V%d.%d' % (d[1], d[0]), name
        if name == 'BHM':
            return '车辆最高允许充电总电压%.1fV' % (u16(d[0], d[1]) / 10.0), name
        if name == 'CRM':
            st = {0x00: '0x00 充电机未识别', 0xAA: '0xAA 充电机已识别'}.get(d[0], '0x%02X' % d[0])
            return 'SPN2560=%s' % st, 'CRM(%02X)' % d[0]
        if name == 'CTS':
            return '时间同步: %02d-%02d-%02d %02d:%02d' % (d[6], d[5], d[4], d[3], d[2]), name
        if name == 'CML':
            return ('最高输出电压%.1fV, 最低输出电压%.1fV, 最大输出电流%.1fA, 最小输出电流%.1fA'
                    % (u16(d[0], d[1]) / 10.0, u16(d[2], d[3]) / 10.0,
                       cur_chg(u16(d[4], d[5])), cur_chg(u16(d[6], d[7])))), name
        if name in ('BRO', 'CRO'):
            st = {0xAA: '0xAA 准备就绪', 0x00: '0x00 未准备好'}.get(d[0], '0x%02X' % d[0])
            return st, '%s(%02X)' % (name, d[0])
        if name == 'BCL':
            mode = {1: '恒压充电', 2: '恒流充电'}.get(d[4], '模式0x%02X' % d[4])
            return ('需求电压%.1fV, 需求电流%.1fA, %s'
                    % (u16(d[0], d[1]) / 10.0, cur_chg(u16(d[2], d[3])), mode)), name
        if name == 'CCS':
            return ('输出电压%.1fV, 输出电流%.1fA, 累计充电时间%dmin'
                    % (u16(d[0], d[1]) / 10.0, cur_chg(u16(d[2], d[3])), u16(d[4], d[5]))), name
        if name == 'BSM':
            return ('最高单体电压编号%d, 最高温度%d℃, 状态位0x%02X%02X'
                    % (d[0], d[1] - 50, d[5], d[6])), name
        if name == 'BDC':
            return ('最高允许放电电流%+.1fA, 最低允许放电电压%.1fV, 最低允许放电SOC=%d%%'
                    % (cur_dis(u16(d[0], d[1])), u16(d[2], d[3]) / 10.0, d[4])), name
        if name == 'BDST':
            r = []
            if d[0] & 3 == 1:
                r.append('接收设备请求放电报文(ERD)超时')
            if (d[0] >> 2) & 3 == 1:
                r.append('接收设备放电控制报文(CCS)超时')
            if d[1] & 3 == 1:
                r.append('收到设备放电中止(EDST)')
            return '车辆放电中止: ' + ('; '.join(r) if r else '正常/车辆主动'), name
        if name == 'EDST':
            r = []
            if d[0] & 3 == 1:
                r.append('接收车辆放电应答报文(BDR)超时')
            if (d[0] >> 2) & 3 == 1:
                r.append('接收车辆放电控制报文(BDC)超时')
            if d[1] & 3 == 1:
                r.append('收到车辆放电中止(BDST)')
            return '设备放电中止: ' + ('; '.join(r) if r else '正常/设备主动'), name
        if name == 'ESD':
            return ('累计放电电量%.1fkWh, 累计放电时长%dmin'
                    % (u16(d[0], d[1]) / 10.0, u16(d[2], d[3]))), name
        if name in ('BST', 'CST'):
            who = '车辆(BMS)' if name == 'BST' else '设备(充电机)'
            return '%s中止: 原因0x%02X 故障0x%02X%02X 错误0x%02X' % (who, d[0], d[1], d[2], d[3]), name
        if name.startswith('BST'):
            return '车辆中止: 原因0x%02X' % d[0], name
        if name == 'BSD':
            return ('中止时SOC=%d%%, 单体最低电压%.2fV, 单体最高电压%.2fV'
                    % (d[0], u16(d[1], d[2]) / 100.0, u16(d[3], d[4]) / 100.0)), name
        if name == 'CSD':
            return ('累计充电时间%dmin, 输出总能量%.1fkWh'
                    % (u16(d[0], d[1]), u16(d[2], d[3]) / 10.0)), name
        if name == 'BEM':
            return '车辆错误报文: 0x%02X%02X%02X%02X' % (d[0], d[1], d[2], d[3]), name
        if name == 'CEM':
            return '设备错误报文: 0x%02X%02X%02X%02X' % (d[0], d[1], d[2], d[3]), name
        if name.startswith('BDR(单帧'):
            return ('单帧BDR: 放电状态bit0-1=%s(01=允许放电), 最大允许放电电流%+.1fA'
                    % (format(d[3] & 3, '02b'), cur_dis(u16(d[4], d[5])))), name
        if name in ('BMV', 'BMT', 'BSP'):
            return {'BMV': '单体电压信息', 'BMT': '单体温度信息',
                    'BSP': '预留报文'}[name], name
        return '', name


# ------------------------------------------------------------
#                        CSV 读取
# ------------------------------------------------------------

def _parse_time(s):
    """支持三种常见格式:
       '2695.973.722'  秒.毫秒.微秒
       '16:36:54.123'  时:分:秒(.毫秒)
       '12.3456'       普通浮点秒
    """
    s = s.strip()
    if not s:
        return 0.0
    try:
        if ':' in s:                      # 时:分:秒(.小数)
            hms = s.split(':')
            sec = float(hms[-1])
            for i, v in enumerate(reversed(hms[:-1]), 1):
                sec += float(v) * (60 ** i)
            return sec
        parts = s.split('.')
        if len(parts) >= 3:               # 秒.毫秒.微秒
            return int(parts[0]) + int(parts[1]) / 1e3 + int(parts[2]) / 1e6
        return float(s)
    except ValueError:
        return 0.0


def _find_col(header, keys, exclude=()):
    """按 keys 的优先级顺序查找列; exclude 中的词命中则跳过该列
    (例: 找"帧数据"列时必须排除"数据长度"列)"""
    hs = [h.replace(' ', '') for h in header]
    for k in keys:
        for i, h in enumerate(hs):
            if any(x in h for x in exclude):
                continue
            if k in h:
                return i
    return -1


def read_csv(path):
    """读取抓包CSV, 返回 (帧列表, 使用的编码)。帧: (序号, 时间秒, ID, 数据HEX串, 字节列表)

    编码依次尝试 GBK/UTF-8/Latin-1: 某种编码虽能解码但表头是乱码(找不到列)时,
    继续尝试下一种, 而不是直接失败。
    """
    errs = []
    for enc in ('utf-8-sig', 'gbk', 'utf-8', 'latin-1'):
        try:
            with open(path, encoding=enc, errors='strict') as f:
                header = next(csv.reader(f))
        except UnicodeDecodeError as e:
            errs.append('%s: 解码失败(%s)' % (enc, e.reason))
            continue
        except StopIteration:
            raise ValueError('文件为空')

        ci_id = _find_col(header, ('帧ID', 'CANID', 'ID'), exclude=('长度',))
        ci_data = _find_col(header, ('帧数据', '数据域', 'Data', '数据'),
                            exclude=('长度', 'DLC', 'Len'))
        ci_time = _find_col(header, ('时间标识', '时间戳', '时间', 'Time'))
        if ci_id < 0 or ci_data < 0:
            errs.append('%s: 表头中找不到"帧ID"/"帧数据"列 → %s' % (enc, header))
            continue  # 可能是编码猜错导致的乱码表头, 换下一种编码再试

        rows = []
        n = 0
        with open(path, encoding=enc, errors='replace') as f:
            rd = csv.reader(f)
            next(rd, None)  # 跳过表头
            for r in rd:
                if len(r) <= max(ci_id, ci_data):
                    continue
                sid = r[ci_id].strip()
                if not sid:
                    continue
                try:
                    mid = int(sid, 16)   # 兼容 "0x1CEC56F4" 与 "1CEC56F4"
                except ValueError:
                    continue
                dh = r[ci_data].strip()
                d = []
                for x in dh.replace(',', ' ').split():
                    try:
                        d.append(int(x, 16))
                    except ValueError:
                        pass
                while len(d) < 8:
                    d.append(0)
                n += 1
                seq = n
                if r and r[0].strip().isdigit():
                    seq = int(r[0].strip())
                t = _parse_time(r[ci_time]) if 0 <= ci_time < len(r) else 0.0
                rows.append((seq, t, mid, dh, d))
        if not rows:
            errs.append('%s: 未解析到有效CAN帧' % enc)
            continue
        return rows, enc
    raise ValueError('无法读取该文件:\n  ' + '\n  '.join(errs))


# ------------------------------------------------------------
#                        Excel 导出
# ------------------------------------------------------------

COLOR_V = '1155CC'   # 蓝: 车辆 -> 设备
COLOR_D = 'C05504'   # 橙: 设备 -> 车辆
COLOR_X = '990000'   # 红: 未识别
FILL_V = 'EEF3F8'
FILL_D = 'FDF1E7'
FILL_X = 'FDEBEB'
ACCENT = '1F4E79'


EXCEL_MAX_ROWS = 1048575          # Excel 单表数据行上限(含表头共 1048576)


def export_xlsx(rows, out_path, veh_name='车辆(BMS)', dev_name='设备(充电机)',
                on_progress=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    truncated = 0
    if len(rows) > EXCEL_MAX_ROWS:
        truncated = len(rows) - EXCEL_MAX_ROWS
        rows = rows[:EXCEL_MAX_ROWS]

    dec = Decoder()
    wb = Workbook()
    ws = wb.active
    ws.title = '报文解析'
    ws.append(['序号', '时间(s)', 'Δt(ms)', '方向', 'CAN ID', '报文', '数据(HEX)', '解析'])
    hf = PatternFill('solid', fgColor=ACCENT)
    hfont = Font(bold=True, color='FFFFFF', size=10)
    for c in ws[1]:
        c.font = hfont
        c.fill = hf

    # 样式对象预先建好复用(每单元格新建会让大文件导出慢十几倍)
    ALIGN = Alignment(vertical='center')
    STYLE = {
        'V': (Font(size=9, color=COLOR_V), Font(size=9, bold=True, color=COLOR_V),
              PatternFill('solid', fgColor=FILL_V)),
        'D': (Font(size=9, color=COLOR_D), Font(size=9, bold=True, color=COLOR_D),
              PatternFill('solid', fgColor=FILL_D)),
        '?': (Font(size=9, color=COLOR_X), Font(size=9, bold=True, color=COLOR_X),
              PatternFill('solid', fgColor=FILL_X)),
    }

    t0 = rows[0][1]
    prev = t0
    events = {}
    counts = {}
    gaps = {}
    last_t = {}
    total = len(rows)

    for i, (seq, tsec, mid, dhex, d) in enumerate(rows):
        name, drc = ID_MAP.get(mid, ('未识别', '?'))
        if name == '未识别':
            txt, label = '未在协议表中的CAN ID', '0x%08X' % mid
        else:
            txt, label = dec.decode(name, drc, d)
        dirs = ('%s→%s' % (veh_name, dev_name)) if drc == 'V' else \
               (('%s→%s' % (dev_name, veh_name)) if drc == 'D' else '?')
        dt = (tsec - prev) * 1000
        prev = tsec
        ws.append([seq, round(tsec - t0, 4), round(dt, 1), dirs,
                   '0x%08X' % mid, label, dhex, txt])
        # 注意: 这里必须按坐标取单元格。用 ws.max_row / ws[行号] 会触发 openpyxl 的
        # 全表扫描(max_row、max_column 都是 O(单元格总数)), 循环里调用会退化成 O(n²),
        # 五万帧要跑几分钟。自行维护行号 + ws.cell() 后是线性的。
        font, font_b, fill = STYLE.get(drc, STYLE['?'])
        r = i + 2
        for col in range(1, 9):
            c = ws.cell(row=r, column=col)
            c.fill = fill
            c.font = font
            c.alignment = ALIGN
        ws.cell(r, 4).font = font_b
        ws.cell(r, 6).font = font_b

        counts[label] = counts.get(label, 0) + 1
        if label not in events:
            events[label] = (round(tsec - t0, 4), seq)
        if label in last_t:
            gaps.setdefault(label, []).append((tsec - last_t[label]) * 1000)
        last_t[label] = tsec
        if on_progress and (i & 0x3FF) == 0:
            on_progress(i, total)

    for i, w in enumerate([6, 10, 8, 18, 13, 14, 26, 72], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws['J1'] = '■ 蓝色 = %s → %s' % (veh_name, dev_name)
    ws['J1'].font = Font(size=9, bold=True, color=COLOR_V)
    ws['K1'] = '■ 橙色 = %s → %s' % (dev_name, veh_name)
    ws['K1'].font = Font(size=9, bold=True, color=COLOR_D)
    ws.column_dimensions['J'].width = 26
    ws.column_dimensions['K'].width = 26
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:H%d' % ws.max_row

    # ---- 关键事件时间线 ----
    ws2 = wb.create_sheet('关键事件时间线')
    ws2.append(['事件(首次出现)', '时间(s)', '抓包序号'])
    for c in ws2[1]:
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = hf
    f9 = Font(size=9)
    for j, (k, (t, s)) in enumerate(sorted(events.items(), key=lambda kv: kv[1][0])):
        ws2.append([k, t, s])
        for col in range(1, 4):
            ws2.cell(row=j + 2, column=col).font = f9
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12
    ws2.freeze_panes = 'A2'

    # ---- 报文统计(含周期) ----
    ws3 = wb.create_sheet('报文统计')
    ws3.append(['报文', '帧数', '平均间隔(ms)', '最小(ms)', '最大(ms)'])
    for c in ws3[1]:
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = hf
    # 按名称排序, 同一报文的 RTS/CTS/DT/ACK 相邻便于比对
    for j, (k, v) in enumerate(sorted(counts.items())):
        g = gaps.get(k, [])
        if g:
            ws3.append([k, v, round(sum(g) / len(g), 1), round(min(g), 1), round(max(g), 1)])
        else:
            ws3.append([k, v, '-', '-', '-'])
        for col in range(1, 6):
            ws3.cell(row=j + 2, column=col).font = f9
    for i, w in enumerate([22, 10, 14, 12, 12], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = 'A2'

    wb.save(out_path)
    return len(rows), len(counts), rows[-1][1] - t0, truncated


# ============================================================
#                         图形界面
# ============================================================

class App(ttk.Frame):
    def __init__(self, master):
        super().__init__(master, padding=10)
        self.pack(fill='both', expand=True)
        self.files = []
        self.q = queue.Queue()
        self.outputs = []
        self._build()
        self.after(120, self._poll)

    # ---------------- 界面 ----------------
    def _build(self):
        top = ttk.LabelFrame(self, text=' 1. 选择抓包文件 (CSV) ', padding=8)
        top.pack(fill='both', expand=True)

        bar = ttk.Frame(top)
        bar.pack(fill='x', pady=(0, 6))
        ttk.Button(bar, text='添加文件…', command=self.add_files, width=12).pack(side='left')
        ttk.Button(bar, text='移除选中', command=self.remove_sel, width=10).pack(side='left', padx=4)
        ttk.Button(bar, text='清空列表', command=self.clear_files, width=10).pack(side='left')
        self.lbl_count = ttk.Label(bar, text='未选择文件', foreground='#666')
        self.lbl_count.pack(side='right')

        wrap = ttk.Frame(top)
        wrap.pack(fill='both', expand=True)
        self.lst = tk.Listbox(wrap, height=7, selectmode='extended',
                              font=('Consolas', 9), activestyle='none')
        sb = ttk.Scrollbar(wrap, orient='vertical', command=self.lst.yview)
        self.lst.configure(yscrollcommand=sb.set)
        self.lst.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')

        opt = ttk.LabelFrame(self, text=' 2. 选项 ', padding=8)
        opt.pack(fill='x', pady=8)
        row1 = ttk.Frame(opt)
        row1.pack(fill='x')
        ttk.Label(row1, text='车辆端名称:').pack(side='left')
        self.var_veh = tk.StringVar(value='车辆(BMS)')
        ttk.Entry(row1, textvariable=self.var_veh, width=14).pack(side='left', padx=(4, 14))
        ttk.Label(row1, text='设备端名称:').pack(side='left')
        self.var_dev = tk.StringVar(value='设备(充电机)')
        ttk.Entry(row1, textvariable=self.var_dev, width=14).pack(side='left', padx=4)
        row2 = ttk.Frame(opt)
        row2.pack(fill='x', pady=(6, 0))
        self.var_open = tk.BooleanVar(value=True)
        ttk.Checkbutton(row2, text='完成后自动打开解析结果', variable=self.var_open).pack(side='left')
        ttk.Label(row2, text='(输出文件保存在原CSV同目录, 文件名加"-解析")',
                  foreground='#666').pack(side='left', padx=8)

        act = ttk.Frame(self)
        act.pack(fill='x')
        self.btn_go = ttk.Button(act, text='开始解析并导出 Excel', command=self.start)
        self.btn_go.pack(side='left', ipady=4, ipadx=10)
        ttk.Button(act, text='打开输出文件夹', command=self.open_folder).pack(side='left', padx=8)
        self.pb = ttk.Progressbar(act, mode='determinate', length=180)
        self.pb.pack(side='right')
        self.var_status = tk.StringVar(value='')
        ttk.Label(act, textvariable=self.var_status, foreground='#666').pack(side='right', padx=8)

        logf = ttk.LabelFrame(self, text=' 3. 运行日志 ', padding=6)
        logf.pack(fill='both', expand=True, pady=(8, 0))
        self.txt = tk.Text(logf, height=11, font=('Consolas', 9), wrap='word',
                           bg='#FBFCFD', relief='flat')
        sb2 = ttk.Scrollbar(logf, orient='vertical', command=self.txt.yview)
        self.txt.configure(yscrollcommand=sb2.set, state='disabled')
        self.txt.pack(side='left', fill='both', expand=True)
        sb2.pack(side='right', fill='y')
        self.txt.tag_config('ok', foreground='#1B7F3B')
        self.txt.tag_config('err', foreground='#B00020')
        self.txt.tag_config('info', foreground='#1155CC')
        self.txt.tag_config('dim', foreground='#777')

        self.log('%s %s  —  支持 GB/T 27930 充电 与 GB/T 18487.4 直流V2L放电报文' % (APP_NAME, APP_VER), 'info')
        self.log('用法: 添加 CAN 分析仪导出的 CSV → 点"开始解析" → 生成彩色 Excel。', 'dim')

    # ---------------- 交互 ----------------
    def log(self, msg, tag=None):
        self.txt.configure(state='normal')
        self.txt.insert('end', msg + '\n', tag or ())
        self.txt.see('end')
        self.txt.configure(state='disabled')

    def add_files(self):
        fs = filedialog.askopenfilenames(
            title='选择 CAN 抓包 CSV 文件',
            filetypes=[('CSV 抓包文件', '*.csv'), ('所有文件', '*.*')])
        for f in fs:
            if f not in self.files:
                self.files.append(f)
                self.lst.insert('end', f)
        self._refresh_count()

    def remove_sel(self):
        for i in reversed(self.lst.curselection()):
            self.lst.delete(i)
            del self.files[i]
        self._refresh_count()

    def clear_files(self):
        self.lst.delete(0, 'end')
        self.files.clear()
        self._refresh_count()

    def _refresh_count(self):
        n = len(self.files)
        self.lbl_count.configure(text='已选择 %d 个文件' % n if n else '未选择文件')

    def open_folder(self):
        target = os.path.dirname(self.outputs[-1]) if self.outputs else (
            os.path.dirname(self.files[-1]) if self.files else os.path.expanduser('~'))
        try:
            os.startfile(target)
        except Exception as e:
            messagebox.showerror(APP_NAME, '无法打开文件夹:\n%s' % e)

    # ---------------- 解析 ----------------
    def start(self):
        if not self.files:
            messagebox.showinfo(APP_NAME, '请先添加要解析的 CSV 文件。')
            return
        self.btn_go.configure(state='disabled')
        self.pb.configure(maximum=len(self.files), value=0)
        veh = self.var_veh.get().strip() or '车辆'
        dev = self.var_dev.get().strip() or '设备'
        threading.Thread(target=self._work, args=(list(self.files), veh, dev),
                         daemon=True).start()

    def _work(self, files, veh, dev):
        for idx, path in enumerate(files, 1):
            name = os.path.basename(path)
            self.q.put(('log', ('[%d/%d] 正在解析 %s' % (idx, len(files), name), 'info')))
            try:
                rows, enc = read_csv(path)
                if len(rows) > 50000:
                    self.q.put(('log', ('    共 %d 帧, 数据量较大, 导出需要一些时间…'
                                        % len(rows), 'dim')))

                def prog(i, total, _n=name):
                    self.q.put(('log2', '正在导出 %s  %d/%d' % (_n, i, total)))

                # 依次尝试: 原目录同名 → 原目录改名(上次的还开着) → 桌面(原目录只读时)
                base = os.path.splitext(path)[0]
                desk = os.path.join(os.path.expanduser('~'), 'Desktop',
                                    os.path.splitext(os.path.basename(path))[0])
                attempts = [(base + '-解析.xlsx', None),
                            (base + '-解析(新).xlsx', '原文件正被打开, 另存为'),
                            (desk + '-解析.xlsx', '原目录无法写入, 已改存到桌面')]
                last_exc = None
                for out, note in attempts:
                    try:
                        n, kinds, dur, cut = export_xlsx(rows, out, veh, dev, prog)
                        if note:
                            self.q.put(('log', ('    %s: %s' % (note, out), 'dim')))
                        last_exc = None
                        break
                    except (PermissionError, OSError) as e:
                        last_exc = e
                if last_exc is not None:
                    raise last_exc
                self.q.put(('log', ('    共 %d 帧 / %d 种报文 / 时长 %.3f 秒 (编码 %s)'
                                    % (n, kinds, dur, enc), 'dim')))
                if cut:
                    self.q.put(('log', ('    注意: 超出 Excel 单表行数上限, 末尾 %d 帧未导出'
                                        % cut, 'err')))
                self.q.put(('log', ('    已导出: %s' % out, 'ok')))
                self.q.put(('out', out))
            except Exception as e:
                self.q.put(('log', ('    失败: %s' % e, 'err')))
                self.q.put(('log', (traceback.format_exc(limit=1).strip(), 'dim')))
            self.q.put(('step', idx))
        self.q.put(('done', None))

    def _poll(self):
        try:
            while True:
                kind, payload = self.q.get_nowait()
                if kind == 'log':
                    self.log(*payload) if isinstance(payload, tuple) else self.log(payload)
                elif kind == 'log2':
                    self.var_status.set(payload)
                elif kind == 'step':
                    self.pb.configure(value=payload)
                elif kind == 'out':
                    self.outputs.append(payload)
                elif kind == 'done':
                    self.btn_go.configure(state='normal')
                    self.var_status.set('')
                    self.log('全部完成。', 'ok')
                    if self.var_open.get() and self.outputs:
                        try:
                            os.startfile(self.outputs[-1])
                        except Exception:
                            pass
        except queue.Empty:
            pass
        self.after(120, self._poll)


def main():
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror(APP_NAME,
                             '缺少 openpyxl 库, 无法导出 Excel。\n\n'
                             '请在命令行执行:\n    pip install openpyxl')
        return
    root = tk.Tk()
    root.title('%s %s' % (APP_NAME, APP_VER))
    root.geometry('880x680')
    root.minsize(760, 560)
    try:
        ttk.Style().theme_use('vista')
    except tk.TclError:
        pass
    App(root)
    root.mainloop()


if __name__ == '__main__':
    main()
